import os
import logging
import re
import json
import time
import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import requests
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from tqdm import tqdm

# ==============================================================================
# Constants & Headers
# ==============================================================================
OUTPUT_DIR = "./output"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Referer": "https://finance.naver.com/"
}

# Global Session to reuse connections.
# Note: requests.Session is NOT fully thread-safe. However, it is used here with
# ThreadPoolExecutor for concurrent GET requests, which works reliably in practice.
# Avoid concurrent writes to session state (e.g., updating headers/cookies) from threads.
session = requests.Session()
session.headers.update(HEADERS)

# ==============================================================================
# Clean Helpers
# ==============================================================================
def clean_int(val_str: str) -> int:
    if not val_str:
        return 0
    clean = str(val_str).replace(",", "").replace("원", "").replace("억", "").replace(" ", "").strip()
    if clean == "N/A" or not clean or clean == "-":
        return 0
    try:
        return int(clean)
    except ValueError:
        return 0

def clean_float(val_str: str) -> float:
    if not val_str:
        return None
    clean = str(val_str).replace(",", "").replace("배", "").replace(" ", "").strip()
    if clean == "N/A" or not clean or clean == "-":
        return None
    try:
        return float(clean)
    except ValueError:
        return None

def clean_float_ratio(val_str: str) -> float:
    if not val_str:
        return None
    clean = str(val_str).replace(",", "").replace("%", "").replace(" ", "").strip()
    if clean == "N/A" or not clean or clean == "-":
        return None
    try:
        return float(clean) / 100.0
    except ValueError:
        return None

# ==============================================================================
# Bulk Data Crawlers (Dividends, Alert Lists)
# ==============================================================================
def fetch_dividend_map() -> dict:
    """Scrapes dividend list page from Naver Finance to build a dict of {code: dividend}."""
    print("Collecting dividend list in bulk...")
    div_map = {}
    page = 1
    
    while True:
        url = f"https://finance.naver.com/sise/dividend_list.naver?page={page}"
        try:
            r = session.get(url, timeout=10)
            r.raise_for_status()
            html = r.content.decode("cp949", errors="ignore")
            soup = BeautifulSoup(html, "html.parser")
            
            # Find the main data table
            table = None
            tables = soup.find_all("table")
            for t in tables:
                if t.find("th") and "종목명" in t.text:
                    table = t
                    break
                    
            if not table:
                break
                
            rows = table.find_all("tr")
            parsed_on_page = 0
            for row in rows:
                a_tag = row.find("a", href=re.compile(r"code=\d{6}"))
                if a_tag:
                    code_match = re.search(r"code=(\d{6})", a_tag.get("href", ""))
                    if code_match:
                        code = code_match.group(1)
                        tds = row.find_all("td")
                        # 0: 종목명, 1: 현재가, 2: 기준월, 3: 배당금
                        if len(tds) > 3:
                            div_val = clean_int(tds[3].text)
                            div_map[code] = div_val
                            parsed_on_page += 1
            
            if parsed_on_page == 0:
                break
            page += 1
        except Exception as e:
            print(f"Error fetching dividend page {page}: {e}")
            break
            
    print(f"Dividend mappings built: {len(div_map)} stocks")
    return div_map

def fetch_alert_sets() -> dict:
    """Fetches codes for management, halt, caution, warning, and risk stocks."""
    print("Collecting stock market alert states...")
    alert_sets = {
        "management": set(),
        "trading_halt": set(),
        "caution": set(),
        "warning": set(),
        "risk": set()
    }
    
    urls = {
        "management": "https://finance.naver.com/sise/management.naver",
        "trading_halt": "https://finance.naver.com/sise/trading_halt.naver",
        "caution": "https://finance.naver.com/sise/investment_alert.naver?type=caution",
        "warning": "https://finance.naver.com/sise/investment_alert.naver?type=warning",
        "risk": "https://finance.naver.com/sise/investment_alert.naver?type=risk"
    }
    
    for key, url in urls.items():
        try:
            r = session.get(url, timeout=10)
            r.raise_for_status()
            html = r.content.decode("cp949", errors="ignore")
            soup = BeautifulSoup(html, "html.parser")
            links = soup.find_all("a", href=re.compile(r"code=\d{6}"))
            for l in links:
                code_match = re.search(r"code=(\d{6})", l.get("href", ""))
                if code_match:
                    alert_sets[key].add(code_match.group(1))
            print(f" - {key}: {len(alert_sets[key])} stocks flagged")
        except Exception as e:
            print(f"Error building alert set {key}: {e}")
            
    return alert_sets

# ==============================================================================
# Basic Market Cap Crawlers (Pre-existing)
# ==============================================================================
def fetch_market_cap_page(sosok: int, page: int) -> list:
    url = f"https://finance.naver.com/sise/sise_market_sum.naver?sosok={sosok}&page={page}"
    results = []
    try:
        r = session.get(url, timeout=10)
        r.raise_for_status()
        html = r.content.decode("cp949", errors="ignore")
        soup = BeautifulSoup(html, "html.parser")
        table = soup.find("table", class_="type_2")
        if not table:
            return []
            
        rows = table.find_all("tr")
        for row in rows:
            a_tag = row.find("a", class_="tltle")
            if not a_tag:
                continue
                
            href = a_tag.get("href", "")
            code_match = re.search(r"code=(\d{6})", href)
            if not code_match:
                continue
                
            code = code_match.group(1)
            name = a_tag.text.strip()
            
            tds = row.find_all("td")
            if len(tds) < 10:
                continue
                
            price = clean_int(tds[2].text)
            market_cap = clean_int(tds[6].text)
            foreign_ratio = clean_float_ratio(tds[8].text)
            
            market_type = "코스피" if sosok == 0 else "코스닥"
            
            results.append({
                "종목코드": code,
                "종목명": name,
                "시장": market_type,
                "현재가": price,
                "시가총액(억)": market_cap,
                "외국인비율": foreign_ratio
            })
    except Exception as e:
        print(f"Error fetching market cap sosok={sosok}, page={page}: {e}")
        
    return results

def fetch_all_stocks() -> pd.DataFrame:
    print("Collecting all KOSPI and KOSDAQ stocks basic information...")
    all_data = []
    tasks = []
    with ThreadPoolExecutor(max_workers=15) as executor:
        for page in range(1, 31):
            tasks.append(executor.submit(fetch_market_cap_page, 0, page))
        for page in range(1, 51):
            tasks.append(executor.submit(fetch_market_cap_page, 1, page))
            
        for future in as_completed(tasks):
            res = future.result()
            if res:
                all_data.extend(res)
                
    df = pd.DataFrame(all_data)
    df = df.drop_duplicates(subset=["종목코드"])
    print(f"Total stocks fetched: {len(df)}")
    return df

def fetch_kospi200_codes() -> set:
    print("Collecting KOSPI 200 constituent codes...")
    kospi200_codes = set()
    
    def fetch_k200_page(page):
        url = f"https://finance.naver.com/sise/entryJongmok.naver?no=028&page={page}"
        codes = []
        try:
            r = session.get(url, timeout=10)
            r.raise_for_status()
            html = r.content.decode("cp949", errors="ignore")
            soup = BeautifulSoup(html, "html.parser")
            links = soup.find_all("a", href=re.compile(r"/item/main\.naver\?code=\d{6}"))
            for link in links:
                code_match = re.search(r"code=(\d{6})", link.get("href", ""))
                if code_match:
                    codes.append(code_match.group(1))
        except Exception as e:
            print(f"Error fetching KOSPI 200 page {page}: {e}")
        return codes

    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = [executor.submit(fetch_k200_page, p) for p in range(1, 21)]
        for future in as_completed(futures):
            kospi200_codes.update(future.result())
            
    print(f"KOSPI 200 codes collected: {len(kospi200_codes)}")
    return kospi200_codes

def fetch_kosdaq150_names() -> set:
    print("Collecting KOSDAQ 150 constituent names...")
    kosdaq150_names = set()
    url = "http://companyinfo.stock.naver.com/v1/company/c1010001.aspx?cmp_cd=229200"
    try:
        r = session.get(url, timeout=10)
        r.raise_for_status()
        html = r.text
        pattern = re.compile(r"var\s+CU_data\s*=\s*(\{.*?\});", re.DOTALL)
        match = pattern.search(html)
        if match:
            json_str = match.group(1)
            data = json.loads(json_str)
            grid_data = data.get("grid_data", [])
            for item in grid_data:
                stk_name = item.get("STK_NM_KOR")
                if stk_name and stk_name != "원화현금":
                    kosdaq150_names.add(stk_name.strip())
    except Exception as e:
        print(f"Error collecting KOSDAQ 150 names: {e}")
        
    print(f"KOSDAQ 150 names collected: {len(kosdaq150_names)}")
    return kosdaq150_names

# ==============================================================================
# Stock Detail Crawling (WICS + Consensus + Highs/Lows)
# ==============================================================================
def fetch_single_stock_detail(code: str) -> dict:
    """Fetches details for a single stock by calling mobile API and WiseFN page."""
    detail = {
        "종목코드": code,
        "KRX업종": "",
        "WICS업종": "",
        "투자의견": None,
        "목표주가": 0,
        "52주최고": 0,
        "52주최저": 0,
        "PER": None,
        "PBR": None,
        "배당수익률": None
    }
    
    # 1. Fetch from Mobile integration API (consensus, ratios, 52-week highs/lows)
    url_api = f"https://m.stock.naver.com/api/stock/{code}/integration"
    max_retries = 3
    for attempt in range(max_retries):
        try:
            r_api = session.get(url_api, timeout=10)
            if r_api.status_code == 200:
                data = r_api.json()
                
                # Consensus Info
                cns = data.get("consensusInfo")
                if cns:
                    detail["투자의견"] = clean_float(cns.get("recommMean"))
                    detail["목표주가"] = clean_int(cns.get("priceTargetMean"))
                    
                # Total Infos (52w high/low, per, pbr, div yield)
                total_infos = data.get("totalInfos", [])
                for info in total_infos:
                    code_key = info.get("code")
                    val = info.get("value")
                    if code_key == "highPriceOf52Weeks":
                        detail["52주최고"] = clean_int(val)
                    elif code_key == "lowPriceOf52Weeks":
                        detail["52주최저"] = clean_int(val)
                    elif code_key == "per":
                        detail["PER"] = clean_float(val)
                    elif code_key == "pbr":
                        detail["PBR"] = clean_float(val)
                    elif code_key == "dividendYieldRatio":
                        detail["배당수익률"] = clean_float_ratio(val)
                break
            elif r_api.status_code == 429:
                time.sleep(2 * (attempt + 1))
            else:
                break
        except Exception as e:
            logging.warning(f'[{code}] detail fetch failed (API): {e}')
            time.sleep(1)
            
    # 2. Fetch from WiseFN page (WICS industry sector & KRX sector)
    url_wise = f"http://companyinfo.stock.naver.com/v1/company/c1010001.aspx?cmp_cd={code}"
    for attempt in range(max_retries):
        try:
            r_wise = session.get(url_wise, timeout=10)
            if r_wise.status_code == 200:
                html = r_wise.text
                
                # Find WICS: 반도체와반도체장비
                wics_match = re.search(r"WICS\s*:\s*([^<]+)", html)
                if wics_match:
                    detail["WICS업종"] = wics_match.group(1).strip()
                    
                # Find KRX Sector (e.g. KOSPI : 코스피 전기·전자 or KOSDAQ : 코스닥 반도체)
                krx_match = re.search(r"(?:KOSPI|KOSDAQ|KONEX)\s*:\s*([^<]+)", html)
                if krx_match:
                    raw_krx = krx_match.group(1).strip()
                    detail["KRX업종"] = re.sub(r"^(코스피|코스닥|코넥스)\s+", "", raw_krx)
                break
            elif r_wise.status_code == 429:
                time.sleep(2 * (attempt + 1))
            else:
                break
        except Exception as e:
            logging.warning(f'[{code}] detail fetch failed (WiseFN): {e}')
            time.sleep(1)
            
    return detail

def fetch_stock_details_parallel(codes: list) -> pd.DataFrame:
    """Fetches details for all stock codes in parallel."""
    print("Collecting stock detailed metrics and WICS sectors in parallel...")
    details = []
    
    with ThreadPoolExecutor(max_workers=20) as executor:
        # Wrap tasks in tqdm progress bar
        futures = {executor.submit(fetch_single_stock_detail, code): code for code in codes}
        for future in tqdm(as_completed(futures), total=len(codes), desc="Details Scraping"):
            try:
                res = future.result()
                if res:
                    details.append(res)
            except Exception as e:
                logging.warning(f'[{futures[future]}] detail fetch failed: {e}')
                
    return pd.DataFrame(details)

# ==============================================================================
# Save & Style Excel Report
# ==============================================================================
def save_to_excel_with_style(df: pd.DataFrame, filepath: str):
    print(f"Saving styled Excel report to: {filepath}...")
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "국내주식 종목현황"
    
    headers = list(df.columns)
    ws.append(headers)
    
    for row in df.itertuples(index=False):
        ws.append(list(row))
        
    font_family = "Malgun Gothic"
    header_font = Font(name=font_family, size=10, bold=True, color="1E293B")
    data_font = Font(name=font_family, size=10, bold=False, color="000000")
    
    header_fill = PatternFill(start_color="EDF2F7", end_color="EDF2F7", fill_type="solid")
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    # Format Headers (Row 1)
    ws.row_dimensions[1].height = 25
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border_all
        
    # Format Data Rows
    # Columns: 
    # 1:종목코드, 2:종목명, 3:유형, 4:KRX업종, 5:WICS업종, 6:현재가, 7:시가총액(억), 
    # 8:외국인비율, 9:투자의견, 10:목표주가, 11:52주최고, 12:52주최저, 13:PER, 14:PBR, 
    # 15:배당수익률, 16:배당금, 17:관리종목
    for r_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[r_idx].height = 20
        
        # Format alignments and number formats per column
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=col_idx)
            cell.font = data_font
            cell.border = border_all
            
            # Alignments
            if col_idx in [1, 3, 4, 5, 17]:
                cell.alignment = align_center
            elif col_idx in [2]:
                cell.alignment = align_left
            else:
                cell.alignment = align_right
                
            # Number formats
            if col_idx == 1:
                cell.number_format = "@" # Text stock code
            elif col_idx in [6, 7, 10, 11, 12, 16]:
                cell.number_format = "#,##0" # Currency/Integer
            elif col_idx in [8, 15]:
                cell.number_format = "0.00%" # Percentages
            elif col_idx in [9, 13, 14]:
                cell.number_format = "0.00" # Ratios
                
    ws.views.sheetView[0].showGridLines = True
    ws.auto_filter.ref = ws.dimensions
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        col_idx = col[0].column
        for cell in col:
            val_str = str(cell.value or '')
            if cell.row > 1:
                if col_idx in [6, 7, 10, 11, 12, 16]:
                    try:
                        val_str = f"{int(cell.value):,}"
                    except (ValueError, TypeError):
                        pass
                elif col_idx in [8, 15]:
                    try:
                        val_str = f"{float(cell.value) * 100:.2f}%"
                    except (ValueError, TypeError):
                        pass
                elif col_idx in [9, 13, 14]:
                    try:
                        val_str = f"{float(cell.value):.2f}"
                    except (ValueError, TypeError):
                        pass
            
            cell_len = 0
            for char in val_str:
                if ord(char) > 127:
                    cell_len += 2
                else:
                    cell_len += 1
            if cell_len > max_len:
                max_len = cell_len
                
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    wb.save(filepath)
    print("Excel file saved successfully!")

# ==============================================================================
# Main Execution
# ==============================================================================
def main():
    start_time = datetime.datetime.now()
    
    # 1. Fetch all stocks basic info
    df_stocks = fetch_all_stocks()
    if df_stocks.empty:
        print("Failed to fetch any stock data.")
        return
        
    # 2. Fetch bulk datasets
    k200_codes = fetch_kospi200_codes()
    k150_names = fetch_kosdaq150_names()
    dividend_map = fetch_dividend_map()
    alert_sets = fetch_alert_sets()
    
    # 3. Fetch detailed metrics & WICS sector (parallel)
    codes_list = list(df_stocks["종목코드"])
    df_details = fetch_stock_details_parallel(codes_list)
    
    # Merge basic stock info with details
    df_merged = pd.merge(df_stocks, df_details, on="종목코드", how="left")
    
    # 4. Process calculations and bulk maps
    print("Post-processing and merging bulk maps...")
    types = []
    dividends = []
    alerts = []
    
    for row in df_merged.itertuples():
        code = row.종목코드
        name = row.종목명
        mkt = row.시장
        
        # A. 유형
        if mkt == "코스피":
            if code in k200_codes:
                types.append("코스피200")
            else:
                types.append("코스피")
        elif mkt == "코스닥":
            if name in k150_names:
                types.append("코스닥150")
            else:
                types.append("코스닥")
        else:
            types.append(mkt)
            
        # B. 배당금
        dividends.append(dividend_map.get(code, 0))
        
        # C. 관리종목 (Alert Flagging)
        flags = []
        if code in alert_sets["management"]:
            flags.append("관리종목")
        if code in alert_sets["trading_halt"]:
            flags.append("거래정지")
        if code in alert_sets["caution"]:
            flags.append("투자주의")
        if code in alert_sets["warning"]:
            flags.append("투자경고")
        if code in alert_sets["risk"]:
            flags.append("투자위험")
            
        alerts.append(", ".join(flags) if flags else "")
        
    df_merged["유형"] = types
    df_merged["KRX업종"] = df_merged["KRX업종"].fillna("")
    df_merged["배당금"] = dividends
    df_merged["관리종목"] = alerts
    
    # 5. Final Columns Selection & Ordering
    final_cols = [
        "종목코드", "종목명", "유형", "KRX업종", "WICS업종", "현재가", 
        "시가총액(억)", "외국인비율", "투자의견", "목표주가", 
        "52주최고", "52주최저", "PER", "PBR", "배당수익률", "배당금", "관리종목"
    ]
    df_final = df_merged[final_cols].copy()
    
    # Sort by market capitalization descending
    df_final = df_final.sort_values(by="시가총액(억)", ascending=False)
    
    # 6. Save Excel
    today_str = datetime.date.today().strftime("%Y%m%d")
    output_path = os.path.join(OUTPUT_DIR, f"Stock_{today_str}.xlsx")
    save_to_excel_with_style(df_final, output_path)
    
    end_time = datetime.datetime.now()
    print(f"Total Elapsed Time: {end_time - start_time}")

if __name__ == "__main__":
    main()
