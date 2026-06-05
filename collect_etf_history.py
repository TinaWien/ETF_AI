"""
국내 상장 ETF 과거 시세 데이터 수집 및 엑셀 출력 스크립트
작성일: 2026-06-05
설명: 국내 상장된 ETF의 과거 시세 데이터(실제가격 및 수정가격)를 수집하여 
      디자인 스타일이 적용된 두 개의 시트(실제가격, 수정주가)를 가진
      엑셀 파일(./output/ETF_YYYYMMDD_Data.xlsx)로 저장합니다.
      수집 소스는 Naver Finance -> KRX (pykrx) -> Yahoo Finance (yfinance) 순으로 
      개별 종목별 독립 폴백(Per-ticker Fallback) 방식을 취하며, 최종 행에 출처를 표기합니다.
"""

import os
import re
import ast
import argparse
import concurrent.futures
from datetime import datetime, timedelta
from typing import Dict, Tuple, Optional, Any, List

import requests
import pandas as pd
from bs4 import BeautifulSoup
from tqdm import tqdm
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import yfinance as yf
from pykrx import stock

# ==============================================================================
# 설정 및 상수 정의
# ==============================================================================
LIST_API_URL = "https://finance.naver.com/api/sise/etfItemList.nhn"
OUTPUT_DIR = "./output"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}

# 엑셀 스타일 디자인 상수
FONT_FAMILY = "Malgun Gothic"
COLOR_HEADER_BG = "EDF2F7"
COLOR_HEADER_TEXT = "1E293B"
COLOR_BORDER = "CBD5E1"

# ==============================================================================
# 날짜 처리 헬퍼 함수
# ==============================================================================
def parse_date_input(date_str: Optional[str], default_str: str) -> str:
    """
    사용자 입력 날짜를 YYYYMMDD 형식으로 정규화합니다.
    """
    if not date_str:
        return default_str
    digits = re.sub(r'\D', '', str(date_str))
    if len(digits) == 8:
        return digits
    return default_str

def to_iso_format(yyyymmdd: str) -> str:
    """
    YYYYMMDD 형식을 YYYY-MM-DD 형식으로 변환합니다.
    """
    return f"{yyyymmdd[:4]}-{yyyymmdd[4:6]}-{yyyymmdd[6:]}"

# ==============================================================================
# 개별 데이터 소스별 수집 함수 (Naver, KRX, Yahoo)
# ==============================================================================
def fetch_from_naver(code: str, timeframe: str, start_date: str, end_date: str, price_type: str) -> Optional[pd.DataFrame]:
    """
    네이버 금융 모바일 차트 API를 사용하여 시세를 수집합니다.
    (네이버는 수정주가 데이터만 제공하므로 실제/수정 요청 모두 동일하게 처리합니다.)
    """
    # timeframe: 'month' 또는 'day'
    url = f"https://m.stock.naver.com/front-api/external/chart/domestic/info?symbol={code}&requestType=1&startTime={start_date}&endTime={end_date}&timeframe={timeframe}"
    try:
        r = requests.get(url, headers=HEADERS, timeout=10)
        if r.status_code != 200:
            return None
        raw_text = r.text.strip()
        if not raw_text or raw_text == "[[]]" or "날짜" not in raw_text:
            return None
        
        # ast.literal_eval을 사용하여 네이버 특유의 작은따옴표 JSON 포맷 파싱
        data = ast.literal_eval(raw_text)
        if len(data) <= 1:
            return None
            
        columns = data[0]
        rows = data[1:]
        df = pd.DataFrame(rows, columns=columns)
        
        # 필요한 열 이름 매핑 및 선택
        df = df.rename(columns={
            '날짜': '날짜',
            '시가': '시가',
            '고가': '고가',
            '저가': '저가',
            '종가': '종가'
        })
        
        # 날짜 포맷팅 YYYYMMDD -> YYYY-MM-DD
        df['날짜'] = df['날짜'].apply(lambda x: f"{x[:4]}-{x[4:6]}-{x[6:]}")
        df['종목코드'] = code
        df['출처'] = 'Naver'
        
        # 수치형 데이터 정리 (정수로 반올림)
        for col in ['시가', '고가', '저가', '종가']:
            df[col] = pd.to_numeric(df[col], errors='coerce').round().fillna(0).astype(int)
            
        df = df[['날짜', '종목코드', '시가', '고가', '저가', '종가', '출처']]
        return df
    except Exception:
        return None

def fetch_from_krx(code: str, timeframe: str, start_date: str, end_date: str, price_type: str) -> Optional[pd.DataFrame]:
    """
    pykrx 라이브러리를 통해 KRX 웹서비스에서 시세를 수집합니다.
    조회 기간이 넓을 경우 발생할 수 있는 데이터 누락을 방지하기 위해 5년 단위로 청킹하여 수집합니다.
    """
    start_dt = datetime.strptime(start_date, "%Y%m%d")
    end_dt = datetime.strptime(end_date, "%Y%m%d")
    
    curr_start = start_dt
    dfs = []
    
    # pykrx의 adjusted 파라미터 매핑
    adjusted_flag = (price_type == 'adjusted')
    
    while curr_start <= end_dt:
        curr_end = curr_start + timedelta(days=5 * 365)
        if curr_end > end_dt:
            curr_end = end_dt
            
        s_str = curr_start.strftime("%Y%m%d")
        e_str = curr_end.strftime("%Y%m%d")
        
        try:
            df_chunk = stock.get_market_ohlcv_by_date(s_str, e_str, code, adjusted=adjusted_flag)
            if df_chunk is not None and not df_chunk.empty:
                # 유효한 행 필터링 (시가와 종가가 0보다 큰 데이터만)
                df_chunk = df_chunk[(df_chunk['시가'] > 0) & (df_chunk['종가'] > 0)]
                if not df_chunk.empty:
                    dfs.append(df_chunk)
        except Exception:
            pass
            
        curr_start = curr_end + timedelta(days=1)
        
    if not dfs:
        return None
        
    df = pd.concat(dfs)
    df = df[~df.index.duplicated(keep='first')]
    
    if df.empty:
        return None
        
    # 월별 데이터 가공이 필요한 경우 Pandas Resample 수행
    if timeframe == 'month':
        df = df.resample('M').agg({
            '시가': 'first',
            '고가': 'max',
            '저가': 'min',
            '종가': 'last'
        })
        df = df.dropna(subset=['시가', '고가', '저가', '종가'])
        
    df = df.reset_index()
    df.columns = ['날짜', '시가', '고가', '저가', '종가']
    df['날짜'] = df['날짜'].dt.strftime('%Y-%m-%d')
    df['종목코드'] = code
    df['출처'] = 'KRX'
    
    # 수치형 데이터 정리 (정수로 반올림)
    for col in ['시가', '고가', '저가', '종가']:
        df[col] = pd.to_numeric(df[col], errors='coerce').round().fillna(0).astype(int)
        
    df = df[['날짜', '종목코드', '시가', '고가', '저가', '종가', '출처']]
    return df

def fetch_from_yahoo(code: str, timeframe: str, start_date: str, end_date: str, price_type: str) -> Optional[pd.DataFrame]:
    """
    yfinance 라이브러리를 사용하여 시세를 수집합니다.
    """
    start_iso = to_iso_format(start_date)
    end_iso = to_iso_format(end_date)
    
    ticker_symbol = f"{code}.KS"
    interval = "1mo" if timeframe == "month" else "1d"
    
    try:
        ticker = yf.Ticker(ticker_symbol)
        df = ticker.history(start=start_iso, end=end_iso, interval=interval, auto_adjust=False)
        if df is None or df.empty:
            return None
            
        df = df.reset_index()
        
        # 수정주가가 필요한 경우 Adj Close 기준으로 OHLC를 재조정
        if price_type == 'adjusted' and 'Adj Close' in df.columns:
            ratio = (df['Adj Close'] / df['Close']).fillna(1.0)
            df['Open'] = df['Open'] * ratio
            df['High'] = df['High'] * ratio
            df['Low'] = df['Low'] * ratio
            df['Close'] = df['Adj Close']
            
        df = df.rename(columns={
            'Date': '날짜',
            'Open': '시가',
            'High': '고가',
            'Low': '저가',
            'Close': '종가'
        })
        
        # 날짜 문자열 정리 및 종목코드/출처 추가
        df['날짜'] = pd.to_datetime(df['날짜']).dt.strftime('%Y-%m-%d')
        df['종목코드'] = code
        df['출처'] = 'Yahoo'
        
        # 수치형 데이터 정리 (정수로 반올림)
        for col in ['시가', '고가', '저가', '종가']:
            df[col] = pd.to_numeric(df[col], errors='coerce').round().fillna(0).astype(int)
            
        df = df[['날짜', '종목코드', '시가', '고가', '저가', '종가', '출처']]
        return df
    except Exception:
        return None

# ==============================================================================
# 단일 종목 수집 통합 함수 (독립 3단계 폴백 루프)
# ==============================================================================
def get_single_ticker_price(code: str, price_type: str, timeframe: str, start_date: str, end_date: str) -> Optional[pd.DataFrame]:
    """
    단일 종목코드에 대해 Naver -> KRX -> Yahoo 순으로 데이터 수집을 시도합니다.
    최종 획득된 데이터프레임에 대해 요청한 날짜 범위로 엄격히 필터링을 수행합니다.
    """
    df = None
    
    # 1단계: 네이버 금융 API 시도
    try:
        df = fetch_from_naver(code, timeframe, start_date, end_date, price_type)
    except Exception:
        df = None
        
    # 2단계: KRX API 시도 (네이버 실패 시)
    if df is None or df.empty:
        try:
            df = fetch_from_krx(code, timeframe, start_date, end_date, price_type)
        except Exception:
            df = None
            
    # 3단계: 야후 파이낸스 API 시도 (네이버 및 KRX 실패 시)
    if df is None or df.empty:
        try:
            df = fetch_from_yahoo(code, timeframe, start_date, end_date, price_type)
        except Exception:
            df = None
            
    # 데이터가 정상 수집되었을 경우 날짜 기준으로 한 번 더 필터링 및 정렬
    if df is not None and not df.empty:
        start_iso = to_iso_format(start_date)
        end_iso = to_iso_format(end_date)
        df = df[(df['날짜'] >= start_iso) & (df['날짜'] <= end_iso)]
        df = df.sort_values(by='날짜').reset_index(drop=True)
        return df
        
    return None

# ==============================================================================
# 다중 종목 병렬 수집 함수
# ==============================================================================
def get_multiple_prices(code_list: List[str], price_type: str, timeframe: str, start_date: str, end_date: str) -> pd.DataFrame:
    """
    여러 종목코드 목록에 대해 ThreadPoolExecutor를 이용하여 병렬로 데이터를 수집 및 병합합니다.
    """
    combined_dfs = []
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=20) as executor:
        # 작업 등록
        future_to_code = {
            executor.submit(get_single_ticker_price, code, price_type, timeframe, start_date, end_date): code
            for code in code_list
        }
        
        # 진행상황 표시
        pbar = tqdm(concurrent.futures.as_completed(future_to_code), total=len(future_to_code), desc=f"수집 중 ({price_type})")
        for future in pbar:
            code = future_to_code[future]
            try:
                df = future.result()
                if df is not None and not df.empty:
                    combined_dfs.append(df)
            except Exception as e:
                print(f"\n[{code}] 에러 발생: {e}")
                
    if not combined_dfs:
        return pd.DataFrame(columns=['날짜', '종목코드', '시가', '고가', '저가', '종가', '출처'])
        
    # 전체 종목 데이터 병합 후 날짜/종목코드 순 정렬
    final_df = pd.concat(combined_dfs, ignore_index=True)
    final_df = final_df.sort_values(by=['날짜', '종목코드']).reset_index(drop=True)
    return final_df

# ==============================================================================
# 엑셀 서식 적용 함수
# ==============================================================================
def apply_excel_styling(workbook: openpyxl.Workbook, sheet_name: str, df: pd.DataFrame) -> None:
    """
    지정된 시트에 맑은 고딕, 테두리, 배경색, 숫자 포맷 등 요구사항 스타일을 적용합니다.
    (대용량 데이터 대응을 위해 Pandas를 활용해 열 너비를 사전 계산하고 최적화된 루프로 셀 서식을 적용합니다.)
    """
    ws = workbook[sheet_name]
    
    # 눈금선 표시 강제 활성화
    ws.views.sheetView[0].showGridLines = True
    
    # 오토필터 활성화
    ws.auto_filter.ref = ws.dimensions
    
    # 스타일 구성 요소 정의
    header_font = Font(name=FONT_FAMILY, size=10, bold=True, color=COLOR_HEADER_TEXT)
    data_font = Font(name=FONT_FAMILY, size=10, color="000000")
    header_fill = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color=COLOR_BORDER),
        right=Side(style='thin', color=COLOR_BORDER),
        top=Side(style='thin', color=COLOR_BORDER),
        bottom=Side(style='thin', color=COLOR_BORDER)
    )
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    # 1. 헤더 스타일링 (1행, 높이 25)
    ws.row_dimensions[1].height = 25
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = align_center
        
    # 기본 행 높이 설정 (개별 행의 height 설정 overhead를 O(1)으로 단축)
    ws.sheet_format.defaultRowHeight = 20
        
    # 2. 데이터 스타일링 및 표시 서식 지정 (O(N) cell access 최적화를 위해 iter_rows 사용)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for col_idx, cell in enumerate(row, start=1):
            cell.font = data_font
            cell.border = thin_border
            
            # 열별 정렬 및 표시 서식 (1:날짜, 2:종목코드, 3:시가, 4:고가, 5:저가, 6:종가, 7:출처)
            if col_idx in [1, 2, 7]:
                cell.alignment = align_center
            elif col_idx in [3, 4, 5, 6]:
                cell.alignment = align_right
                cell.number_format = '#,##0'  # 천 단위 쉼표 정수
                
            # 종목코드는 문자열 처리
            if col_idx == 2:
                cell.number_format = '@'
                
    # 3. 열 너비 자동 조정 (Pandas 벡터화 연산으로 O(N) 문자열 길이 계산을 O(1)에 가깝게 단축)
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        # 각 셀값의 문자열 길이를 한글 가중치(ord > 127 이면 2, 아니면 1)를 고려해 계산
        if not df.empty:
            col_len = df[col_name].fillna("").astype(str).apply(
                lambda x: sum(2 if ord(c) > 127 else 1 for c in x)
            ).max()
        else:
            col_len = 0
            
        header_len = sum(2 if ord(c) > 127 else 1 for c in str(col_name))
        max_len = max(col_len, header_len)
        
        # 가독성을 위해 마진 4 추가 및 최소 너비 12 설정
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

# ==============================================================================
# 메인 실행 프로세스
# ==============================================================================
def main() -> None:
    # 인자 정의
    parser = argparse.ArgumentParser(description="ETF 과거 시세 데이터 수집 스크립트")
    parser.add_argument("--code", type=str, default=None, help="특정 ETF 종목코드 (예: 069500). 지정하지 않으면 전체 ETF 수집")
    parser.add_argument("--start", type=str, default=None, help="시작 날짜 (YYYYMMDD 형식, 미지정 시 최초 상장일)")
    parser.add_argument("--end", type=str, default=None, help="종료 날짜 (YYYYMMDD 형식, 미지정 시 현재 날짜)")
    parser.add_argument("--timeframe", type=str, default="month", choices=["month", "day"], help="수집 주기 (month 또는 day, 기본값: month)")
    parser.add_argument("--output", type=str, default=None, help="출력 엑셀 파일명 (기본값: ETF_날짜_Data.xlsx)")
    
    args = parser.parse_args()
    
    # 날짜 정규화 및 기본값 처리
    today_str = datetime.today().strftime("%Y%m%d")
    start_date = parse_date_input(args.start, "19900101")
    end_date = parse_date_input(args.end, today_str)
    
    print(f"==================================================================")
    print(f" ETF 과거 시세 데이터 수집 시작")
    print(f" - 조회 기간: {to_iso_format(start_date)} ~ {to_iso_format(end_date)}")
    print(f" - 조회 주기: {args.timeframe}")
    print(f"==================================================================")
    
    # 1. 수집 대상 ETF 리스트 결정
    if args.code:
        print(f"단일 종목 코드 [{args.code}] 조회를 준비합니다.")
        code_list = [args.code]
    else:
        print("네이버 금융 API에서 전체 ETF 목록을 가져옵니다...")
        try:
            res = requests.get(LIST_API_URL, headers=HEADERS, timeout=10)
            res.raise_for_status()
            data = res.json()
            etf_items = data['result']['etfItemList']
            code_list = [item['itemcode'] for item in etf_items]
            print(f"총 {len(code_list)}개의 ETF 종목이 확인되었습니다.")
        except Exception as e:
            print(f"전체 ETF 목록 조회 실패: {e}")
            return
            
    # 2. 데이터 수집 진행 (실제가격 및 수정가격)
    print("\n[단계 1] 실제가격(Unadjusted) 데이터를 수집합니다...")
    df_actual = get_multiple_prices(code_list, "actual", args.timeframe, start_date, end_date)
    
    print("\n[단계 2] 수정주가(Adjusted) 데이터를 수집합니다...")
    df_adjusted = get_multiple_prices(code_list, "adjusted", args.timeframe, start_date, end_date)
    
    # 3. 엑셀 파일 생성 및 저장
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    if args.output:
        excel_path = os.path.join(OUTPUT_DIR, args.output)
    else:
        excel_path = os.path.join(OUTPUT_DIR, f"ETF_{today_str}_Data.xlsx")
        
    print(f"\n[단계 3] 엑셀 파일 생성 및 서식 지정 중: {excel_path}")
    
    try:
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df_actual.to_excel(writer, sheet_name="실제가격", index=False)
            df_adjusted.to_excel(writer, sheet_name="수정주가", index=False)
            
            workbook = writer.book
            
            # 각 시트에 스타일 적용
            apply_excel_styling(workbook, "실제가격", df_actual)
            apply_excel_styling(workbook, "수정주가", df_adjusted)
            
        print("\n과거 시세 수집 및 엑셀 리포트 생성이 완료되었습니다!")
        print(f" - 생성 완료: {os.path.abspath(excel_path)}")
    except Exception as e:
        print(f"\n엑셀 파일 저장 실패: {e}")

if __name__ == "__main__":
    main()
