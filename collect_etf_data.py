"""
국내 상장 ETF 정보 수집 및 엑셀 출력 스크립트
작성일: 2026-06-05
설명: 네이버 금융 API와 WiseFN 상세 페이지 크롤링을 활용하여 국내 상장된 전체 ETF 정보를 수집하고,
      디자인 스타일이 적용된 단일 시트 엑셀 파일(./output/ETF_YYYYMMDD.xlsx)로 저장합니다.
"""

import os
import re
import json
import concurrent.futures
from datetime import datetime
from typing import Dict, Tuple, Optional, Any, List

import requests
import pandas as pd
from tqdm import tqdm
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import yfinance as yf

# ==============================================================================
# 설정 및 상수 정의
# ==============================================================================
LIST_API_URL = "https://finance.naver.com/api/sise/etfItemList.nhn"
DETAIL_PAGE_URL = "https://finance.naver.com/item/main.naver?code={code}"
OUTPUT_DIR = "./output"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}

# 엑셀 스타일 디자인 상수
FONT_FAMILY = "Malgun Gothic"
COLOR_HEADER_BG = "EDF2F7"
COLOR_HEADER_TEXT = "1E293B"
COLOR_BORDER = "CBD5E1"


# ==============================================================================
# 데이터 정제 및 파싱 헬퍼 함수
# ==============================================================================
def parse_categories(etf_type_str: str, tab_code: int, item_name: str) -> Tuple[str, str, str]:
    """
    네이버 탭 코드와 상세 유형 텍스트, 종목명을 분석하여 대/중/소분류 유형을 매핑합니다.
    """
    # 1. 대분류 결정 ("국내", "해외", "기타")
    large = "기타"
    if tab_code in [1, 2, 3]:
        large = "국내"
    elif tab_code == 4:
        large = "해외"
    elif etf_type_str:
        if "국내" in etf_type_str:
            large = "국내"
        elif "해외" in etf_type_str:
            large = "해외"
            
    # 2. 중분류 결정 ("시장지수", "업종/테마", "파생", "원자재", "채권", "기타")
    medium = "기타"
    if tab_code == 1:
        medium = "시장지수"
    elif tab_code == 2:
        medium = "업종/테마"
    elif tab_code == 3:
        medium = "파생"
    elif tab_code == 4:
        if etf_type_str and any(k in etf_type_str for k in ["시장대표", "대표지수", "대표"]):
            medium = "시장지수"
        else:
            medium = "업종/테마"
    elif tab_code == 5:
        medium = "원자재"
    elif tab_code == 6:
        medium = "채권"
    elif tab_code == 7:
        medium = "기타"
        
    # 3. 소분류 결정 (종목명 내 특정 키워드 조합)
    keywords = ["TR", "액티브", "합성", "(H)", "인버스", "레버리지"]
    matched = [kw for kw in keywords if kw in item_name]
    small = ", ".join(matched) if matched else ""
    
    return large, medium, small


def parse_date(date_str: str) -> str:
    """
    한글 형식의 날짜("2002년 10월 14일")를 ISO 형식("2002-10-14")으로 변환합니다.
    """
    if not date_str or not isinstance(date_str, str):
        return ""
    match = re.search(r'(\d{4})\s*년\s*(\d{1,2})\s*월\s*(\d{1,2})\s*일', date_str)
    if match:
        year, month, day = match.group(1), match.group(2), match.group(3)
        return f"{year}-{int(month):02d}-{int(day):02d}"
    return date_str


def parse_json_var(var_name: str, html_text: str) -> Dict[str, Any]:
    """
    HTML 본문 내 선언된 JavaScript JSON 객체 변수를 정규식으로 안전하게 추출하여 딕셔너리로 반환합니다.
    예: var status_data = { ... }; -> {'...': '...'}
    """
    pattern = rf'var\s+{var_name}\s*=\s*({{.*?}});'
    match = re.search(pattern, html_text, re.DOTALL)
    if match:
        try:
            return json.loads(match.group(1).strip())
        except Exception:
            pass
    return {}


# ==============================================================================
# 웹 스크래퍼 함수
# ==============================================================================
def fetch_etf_detail(code: str, tab_code: int) -> Tuple[str, Optional[Dict[str, Any]]]:
    """
    개별 ETF 상세 정보(기초지수, 유형, 상장일, 보수, 자산운용사, 52주 최고/최저, 수익률, 최근분배금, 지급기준일)를
    WiseFN 상세 페이지 및 Yahoo Finance를 통해 추출합니다.
    """
    url = f"http://companyinfo.stock.naver.com/v1/company/c1010001.aspx?cmp_cd={code}"
    try:
        r = requests.get(url, headers=HEADERS, timeout=10)
        if r.status_code != 200:
            return code, None
            
        text = r.text
        
        # WiseFN 페이지 내 자바스크립트 객체 파싱
        status_data = parse_json_var("status_data", text)
        product_data = parse_json_var("product_summary_data", text)
        summary_data = parse_json_var("summary_data", text)
        
        # 기본 정보 추출
        index_name = product_data.get("BASE_IDX_NM_KOR", "")
        etf_type_str = summary_data.get("ETF_TYP_SVC_NM", product_data.get("FUND_TYP", ""))
        listing_date = product_data.get("LIST_DT", "")
        
        # 수수료 파싱 (TOT_PAY는 연보수(%) 형식 예: '0.150', 천 단위 쉼표 제거 대응)
        tot_pay = product_data.get("TOT_PAY", "")
        fee = None
        if tot_pay:
            try:
                fee = float(tot_pay.replace(",", "")) / 100.0
            except ValueError:
                pass
                
        manager = product_data.get("ISSUE_NM_KOR", "")
        
        # 52주 최고/최저가 추출 (천 단위 쉼표 제거)
        high_52w = status_data.get("YR_HIGH", "").replace(",", "")
        low_52w = status_data.get("YR_LOW", "").replace(",", "")
        high_52w_val = int(high_52w) if high_52w else None
        low_52w_val = int(low_52w) if low_52w else None
        
        # 수익률 파싱 (3/6/12개월, 레버리지 종목 등의 천 단위 쉼표 제거 대응)
        ern3 = status_data.get("ERN3", "")
        ern6 = status_data.get("ERN6", "")
        ern12 = status_data.get("ERN12", "")
        
        val_3m = float(ern3.replace(",", "")) / 100.0 if ern3 else None
        val_6m = float(ern6.replace(",", "")) / 100.0 if ern6 else None
        val_12m = float(ern12.replace(",", "")) / 100.0 if ern12 else None
        
        # 지급기준일
        div_base_dt = product_data.get("DIV_BASE_DT", "")
        
        # 최근 분배금 (yfinance 연동)
        recent_div = 0
        
        # 지급기준일 정보가 유효하고, 배당지급이 없거나 제한된 종목이 아닌 경우에만 yfinance 조회 시도
        has_dividend = div_base_dt and not any(k in div_base_dt for k in ["지급하지", "없음", "미지급"])
        
        if has_dividend:
            try:
                ticker = yf.Ticker(f"{code}.KS")
                divs = ticker.dividends
                if not divs.empty:
                    recent_div = int(divs.iloc[-1])
            except Exception:
                recent_div = 0
                
        return code, {
            "index_name": index_name,
            "etf_type_str": etf_type_str,
            "listing_date": listing_date,
            "fee": fee,
            "manager": manager,
            "high_52w": high_52w_val,
            "low_52w": low_52w_val,
            "return_3m": val_3m,
            "return_6m": val_6m,
            "return_12m": val_12m,
            "div_base_dt": div_base_dt if div_base_dt else "없음",
            "recent_div": recent_div
        }
    except Exception:
        return code, None


# ==============================================================================
# 메인 프로세스
# ==============================================================================
def main() -> None:
    print("1. 네이버 금융 ETF 목록 조회 중...")
    try:
        res = requests.get(LIST_API_URL, headers=HEADERS, timeout=10)
        res.raise_for_status()
        data = res.json()
        etf_items = data['result']['etfItemList']
    except Exception as e:
        print(f"API 조회 실패: {e}")
        return
        
    print(f"총 {len(etf_items)}개의 ETF가 조회되었습니다.")
    
    print("\n2. 개별 ETF 상세 정보 수집 중 (병렬 처리)...")
    details: Dict[str, Dict[str, Any]] = {}
    
    # ThreadPoolExecutor를 사용한 멀티스레딩 병렬 수집 (Yahoo 과도한 차단 방지를 위해 workers=15 지정)
    with concurrent.futures.ThreadPoolExecutor(max_workers=15) as executor:
        futures = {
            executor.submit(fetch_etf_detail, item['itemcode'], item.get('etfTabCode', 7)): item['itemcode']
            for item in etf_items
        }
        
        for future in tqdm(concurrent.futures.as_completed(futures), total=len(futures)):
            code, result = future.result()
            if result:
                details[code] = result
                
    # 3. 데이터 병합 및 분류 매핑
    print("\n3. 데이터 병합 및 분류 매핑 중...")
    combined_data = []
    
    for item in etf_items:
        code = item['itemcode']
        name = item['itemname']
        tab_code = item.get('etfTabCode', 7)
        market_cap = item.get('marketSum', 0)
        volume = item.get('quant', 0)
        
        # 크롤링 정보가 없는 경우 기본값 처리
        detail = details.get(code, {
            "index_name": "",
            "etf_type_str": "",
            "listing_date": "",
            "fee": None,
            "manager": "",
            "high_52w": None,
            "low_52w": None,
            "return_3m": None,
            "return_6m": None,
            "return_12m": None,
            "div_base_dt": "없음",
            "recent_div": 0
        })
        
        large_cat, med_cat, small_cat = parse_categories(detail['etf_type_str'], tab_code, name)
        
        combined_data.append({
            "종목코드": code,
            "종목명": name,
            "자산운용사": detail['manager'],
            "대분류유형": large_cat,
            "중분류유형": med_cat,
            "소분류유형": small_cat,
            "기초지수명": detail['index_name'],
            "상장일": detail['listing_date'],
            "시가총액(억원)": market_cap,
            "거래량": volume,
            "수수료": detail['fee'],
            "지급기준일": detail.get('div_base_dt', '없음'),
            "최근분배금(원)": detail.get('recent_div', 0),
            "수익률(3개월)": detail.get('return_3m'),
            "수익률(6개월)": detail.get('return_6m'),
            "수익률(12개월)": detail.get('return_12m'),
            "52주최고": detail.get('high_52w'),
            "52주최저": detail.get('low_52w')
        })
        
    df = pd.DataFrame(combined_data)
    
    # 4. 엑셀 파일 생성 및 스타일링
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    now_str = datetime.now().strftime("%Y%m%d")
    excel_path = os.path.join(OUTPUT_DIR, f"ETF_{now_str}.xlsx")
    print(f"엑셀 파일 저장 및 서식 적용 중: {excel_path}")
    
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="Sheet1", index=False)
        workbook = writer.book
        ws = workbook["Sheet1"]
        
        # 눈금선 강제 표시 및 오토필터 지정
        ws.views.sheetView[0].showGridLines = True
        ws.auto_filter.ref = ws.dimensions
        
        # 스타일 구성 요소 정의
        header_font = Font(name=FONT_FAMILY, size=10, bold=True, color=COLOR_HEADER_TEXT)
        data_font = Font(name=FONT_FAMILY, size=10, color="000000")
        header_fill = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color=COLOR_BORDER),
            right=Side(style='thin', color=COLOR_BORDER),
            top=Side(style='thin', color=COLOR_BORDER),
            bottom=Side(style='thin', color=COLOR_BORDER)
        )
        
        align_center = Alignment(horizontal='center', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center')
        align_right = Alignment(horizontal='right', vertical='center')
        
        # 1) 헤더 스타일링 (1행, 높이 25)
        ws.row_dimensions[1].height = 25
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = align_center
            
        # 2) 데이터 행 높이 및 스타일링 적용 (O(N) 최적화 루프)
        ws.sheet_format.defaultRowHeight = 20
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for col_idx, cell in enumerate(row, start=1):
                cell.font = data_font
                cell.border = thin_border
                
                # 열별 정렬 및 표시 방식 서식 지정
                # 1:종목코드, 2:종목명, 3:자산운용사, 4:대분류유형, 5:중분류유형, 6:소분류유형, 7:기초지수명, 8:상장일, 9:시가총액(억원), 10:거래량, 11:수수료, 12:지급기준일, 13:최근분배금(원), 14:수익률(3개월), 15:수익률(6개월), 16:수익률(12개월), 17:52주최고, 18:52주최저
                if col_idx in [1, 4, 5, 6, 8, 12]:
                    cell.alignment = align_center
                elif col_idx in [2, 3, 7]:
                    cell.alignment = align_left
                elif col_idx in [9, 10, 11, 13, 14, 15, 16, 17, 18]:
                    cell.alignment = align_right
                    
                # 숫자 포맷 지정
                if col_idx in [9, 10, 13, 17, 18]:
                    cell.number_format = '#,##0'  # 천 단위 쉼표 정수
                elif col_idx == 11:
                    if isinstance(cell.value, (float, int)):
                        cell.number_format = '0.000%'  # 소수점 3자리 백분율
                    else:
                        cell.alignment = align_center
                elif col_idx in [14, 15, 16]:
                    if isinstance(cell.value, (float, int)):
                        cell.number_format = '0.00%'  # 소수점 2자리 백분율
                    else:
                        cell.alignment = align_center
                elif col_idx == 1:
                    cell.number_format = '@'
                            
        # 3) 열 너비 자동 조정 (Pandas 벡터화 연산으로 O(N) 문자열 길이 계산을 O(1) 수준으로 단축)
        for col_idx, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_idx)
            
            # 각 셀값의 문자열 길이를 한글 가중치(ord > 127 이면 2, 아니면 1)를 고려해 계산
            if not df.empty:
                col_len = df[col_name].fillna("").astype(str).apply(
                    lambda x: sum(2 if ord(c) > 127 else 1 for c in x)
                ).max()
            else:
                col_len = 0
                
            header_len = sum(2 if ord(c) > 127 else 1 for c in str(col_name))
            max_len = max(col_len, header_len)
            
            # 가독성을 위해 마진 4 추가 및 최소 너비 12 설정
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    print("\n수집 및 엑셀 생성이 완료되었습니다!")


if __name__ == "__main__":
    main()
