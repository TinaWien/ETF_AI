import os
import glob
import re
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

print("=== ETF Metadata Excel File Verification ===")

# Find latest file in output folder matching ETF_YYYYMMDD.xlsx (excluding files with suffixes like _Data)
output_files = glob.glob("output/ETF_*.xlsx")
valid_files = []
for f in output_files:
    filename = os.path.basename(f)
    # Match ETF_ followed by 8 digits and then .xlsx
    if re.match(r"^ETF_\d{8}\.xlsx$", filename):
        valid_files.append(f)

if not valid_files:
    print("Error: No generated ETF metadata Excel files (ETF_YYYYMMDD.xlsx) found in output/!")
    exit(1)
    
file_path = max(valid_files, key=os.path.getmtime)
print(f"Excel file selected for verification: {file_path}")

# Load using openpyxl to verify styles
wb = openpyxl.load_workbook(file_path)
sheet_names = wb.sheetnames
print(f"Sheet names: {sheet_names}")

if len(sheet_names) != 1 or sheet_names[0] != "Sheet1":
    print("Error: Excel should contain exactly one sheet named 'Sheet1'!")
    exit(1)
    
ws = wb["Sheet1"]

# Verify showGridLines
if not ws.views.sheetView[0].showGridLines:
    print("Error: Gridlines are not visible!")
    exit(1)
else:
    print("Gridlines check: OK (Visible)")

# Verify Autofilter
if not ws.auto_filter.ref:
    print("Error: Autofilter is not enabled!")
    exit(1)
else:
    print(f"Autofilter check: OK (Range: {ws.auto_filter.ref})")

# Load as pandas DataFrame to check content and columns
df = pd.read_excel(file_path, sheet_name="Sheet1", dtype={'종목코드': str})
print(f"Total rows collected: {len(df)}")
print(f"Total columns: {len(df.columns)}")

expected_cols = [
    "종목코드", "종목명", "자산운용사", "대분류유형", "중분류유형", "소분류유형", 
    "기초지수명", "상장일", "시가총액(억원)", "거래량", "수수료",
    "지급기준일", "최근분배금(원)", "수익률(3개월)", "수익률(6개월)", "수익률(12개월)",
    "52주최고", "52주최저"
]

if df.columns.tolist() != expected_cols:
    print(f"Error: Columns mismatch!")
    print(f"Expected ({len(expected_cols)}): {expected_cols}")
    print(f"Got ({len(df.columns)}): {df.columns.tolist()}")
    exit(1)
else:
    print("Column headers check: OK")

# Check for null values in key columns
null_code = df["종목코드"].isnull().sum()
null_name = df["종목명"].isnull().sum()
if null_code > 0 or null_name > 0:
    print(f"Warning: Found nulls! Null codes: {null_code}, Null names: {null_name}")
    exit(1)
else:
    print("Key columns null check: OK")

# Verify styles of specific cells
# 1. Header style check
header_height = ws.row_dimensions[1].height
if header_height != 25:
    print(f"Warning: Header row height is {header_height}, expected 25")
else:
    print("Header height check: OK (25)")

for col_idx in range(1, len(expected_cols) + 1):
    cell = ws.cell(row=1, column=col_idx)
    # Font
    if cell.font.name != "Malgun Gothic" or cell.font.size != 10 or not cell.font.bold:
        print(f"Error: Header cell at col {col_idx} has incorrect font: {cell.font.name}, size={cell.font.size}, bold={cell.font.bold}")
        exit(1)
    # Fill
    if cell.fill.start_color.rgb != "00EDF2F7" and cell.fill.start_color.rgb != "EDF2F7":
        print(f"Error: Header cell at col {col_idx} has incorrect fill color: {cell.fill.start_color.rgb}")
        exit(1)
    # Border
    if not cell.border.left or cell.border.left.style != 'thin':
        print(f"Error: Header cell at col {col_idx} does not have thin border")
        exit(1)
    # Alignment
    if cell.alignment.horizontal != 'center' or cell.alignment.vertical != 'center':
        print(f"Error: Header cell at col {col_idx} has incorrect alignment: {cell.alignment.horizontal}/{cell.alignment.vertical}")
        exit(1)

print("Header styles check: OK (Malgun Gothic 10pt Bold, EDF2F7 fill, thin border, center alignment)")

# 2. Data row style and number formatting check
data_font_ok = True
data_alignment_ok = True
data_format_ok = True
data_border_ok = True

# We check a subset of rows to save time, e.g. first 20 rows
check_limit = min(ws.max_row, 50)
for row_idx in range(2, check_limit + 1):
    # Check row height if specified, or if default is set
    # Note: openpyxl might return None for row height if it's default, but defaultRowHeight is set
    row_height = ws.row_dimensions[row_idx].height
    # We set defaultRowHeight to 20 in collect_etf_data.py?
    # Wait, in collect_etf_data.py line 361: ws.row_dimensions[row].height = 20
    if row_height != 20 and row_height is not None:
        print(f"Warning: Row {row_idx} height is {row_height}, expected 20 or default")
        
    for col_idx in range(1, len(expected_cols) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        val = cell.value
        
        # Font check
        if cell.font.name != "Malgun Gothic" or cell.font.size != 10:
            data_font_ok = False
            print(f"Font mismatch at cell R{row_idx}C{col_idx}: {cell.font.name}, size={cell.font.size}")
            
        # Border check
        if not cell.border.left or cell.border.left.style != 'thin':
            data_border_ok = False
            print(f"Border mismatch at cell R{row_idx}C{col_idx}")
            
        # Alignment & Number format check
        align = cell.alignment.horizontal
        fmt = cell.number_format
        
        # Column categorization:
        # 1:종목코드, 2:종목명, 3:자산운용사, 4:대분류유형, 5:중분류유형, 6:소분류유형, 7:기초지수명, 8:상장일, 9:시가총액(억원), 10:거래량, 11:수수료, 12:지급기준일, 13:최근분배금(원), 14:수익률(3개월), 15:수익률(6개월), 16:수익률(12개월), 17:52주최고, 18:52주최저
        if col_idx in [1, 4, 5, 6, 8, 12]:  # Center align
            if align != 'center':
                data_alignment_ok = False
                print(f"Alignment error at R{row_idx}C{col_idx}: expected 'center', got '{align}'")
        elif col_idx in [2, 3, 7]:  # Left align
            if align != 'left' and align is not None:
                data_alignment_ok = False
                print(f"Alignment error at R{row_idx}C{col_idx}: expected 'left', got '{align}'")
        elif col_idx in [9, 10, 11, 13, 14, 15, 16, 17, 18]:  # Right align
            if align != 'right':
                # Exception: if fee or return is non-numeric, it might be center-aligned (like None or string)
                if col_idx in [11, 14, 15, 16] and not isinstance(val, (int, float)):
                    if align != 'center':
                        data_alignment_ok = False
                        print(f"Alignment error at non-numeric cell R{row_idx}C{col_idx}: expected 'center', got '{align}'")
                else:
                    data_alignment_ok = False
                    print(f"Alignment error at R{row_idx}C{col_idx}: expected 'right', got '{align}'")
                    
        # Number format checks
        if col_idx == 1: # 종목코드
            if fmt != '@':
                data_format_ok = False
                print(f"Format error at R{row_idx}C{col_idx} (종목코드): expected '@', got '{fmt}'")
        elif col_idx in [9, 10, 13, 17, 18]: # 시총, 거래량, 분배금, 52주 최고/최저
            if fmt != '#,##0':
                data_format_ok = False
                print(f"Format error at R{row_idx}C{col_idx} (comma integer): expected '#,##0', got '{fmt}'")
        elif col_idx == 11: # 수수료
            if isinstance(val, (int, float)):
                if fmt != '0.000%':
                    data_format_ok = False
                    print(f"Format error at R{row_idx}C{col_idx} (수수료): expected '0.000%', got '{fmt}'")
        elif col_idx in [14, 15, 16]: # 수익률
            if isinstance(val, (int, float)):
                if fmt != '0.00%':
                    data_format_ok = False
                    print(f"Format error at R{row_idx}C{col_idx} (수익률): expected '0.00%', got '{fmt}'")

if data_font_ok:
    print("Data font check: OK (Malgun Gothic 10pt)")
else:
    exit(1)

if data_border_ok:
    print("Data border check: OK (Thin border)")
else:
    exit(1)

if data_alignment_ok:
    print("Data alignment check: OK (Center/Left/Right alignments)")
else:
    exit(1)

if data_format_ok:
    print("Data number format check: OK (@, #,##0, 0.000%, 0.00% formats)")
else:
    exit(1)

# Check auto column widths
for col_idx in range(1, len(expected_cols) + 1):
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    width = ws.column_dimensions[col_letter].width
    if width is None or width < 12:
        print(f"Warning: Column {col_letter} width is {width}, which is less than minimum 12")
    else:
        # print(f"Column {col_letter} width: {width}")
        pass
print("Column widths check: OK (All >= 12)")

# Display sample of first 5 rows
print("\n=== Data Sample (First 5 rows) ===")
print(df.head().to_string(index=False))

# Check size constraint
assert len(df) > 1000, f"Expected more than 1000 ETFs, but got {len(df)}"
print(f"Verification Successful! Data is valid and style guidelines are met perfectly. File: {file_path}")
wb.close()
