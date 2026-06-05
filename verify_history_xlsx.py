import os
import subprocess
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

TEST_OUTPUT = "ETF_test_Data.xlsx"
TEST_PATH = os.path.join("./output", TEST_OUTPUT)

def run_test_collection():
    print("Running collect_etf_history.py for test ticker 069500 (2023)...")
    # Clean previous output if exists
    if os.path.exists(TEST_PATH):
        os.remove(TEST_PATH)
        
    cmd = [
        "python3", "collect_etf_history.py",
        "--code", "069500",
        "--start", "20230101",
        "--end", "20231231",
        "--output", TEST_OUTPUT
    ]
    res = subprocess.run(cmd, capture_output=True, text=True)
    print("STDOUT:")
    print(res.stdout)
    if res.stderr:
        print("STDERR:")
        print(res.stderr)
        
    if not os.path.exists(TEST_PATH):
        raise FileNotFoundError(f"Test output excel file not found at: {TEST_PATH}")
    print("Excel file successfully generated.")

def verify_excel_file():
    print("\nVerifying Excel properties...")
    wb = openpyxl.load_workbook(TEST_PATH)
    
    # 1. Sheets check
    sheets = wb.sheetnames
    print(f"Sheet names found: {sheets}")
    assert "실제가격" in sheets, "실제가격 missing!"
    assert "수정주가" in sheets, "수정주가 missing!"
    
    for name in ["실제가격", "수정주가"]:
        ws = wb[name]
        print(f"\n--- Checking {name} ({'Actual' if name == '실제가격' else 'Adjusted'}) ---")
        
        # 2. Gridlines check
        assert ws.views.sheetView[0].showGridLines, f"Gridlines not showing on {name}!"
        print("Gridlines are successfully forced visible.")
        
        # 3. Columns check
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        expected_headers = ['날짜', '종목코드', '시가', '고가', '저가', '종가', '출처']
        print(f"Headers: {headers}")
        assert headers == expected_headers, f"Headers mismatch! Found: {headers}, expected: {expected_headers}"
        
        # 4. Rows check
        rows_count = ws.max_row
        print(f"Total rows (including header): {rows_count}")
        assert rows_count > 1, f"No data rows found in {name}!"
        
        # 5. Styling Check (Header)
        header_cell = ws.cell(row=1, column=1)
        font = header_cell.font
        fill = header_cell.fill
        align = header_cell.alignment
        
        print("Checking header cell (A1) style:")
        print(f" - Font name: {font.name}")
        print(f" - Font size: {font.size}")
        print(f" - Font bold: {font.bold}")
        print(f" - Fill color: {fill.start_color.rgb}")
        print(f" - Alignment: {align.horizontal}")
        
        assert font.name == "Malgun Gothic", f"Incorrect font: {font.name}"
        assert font.size == 10, f"Incorrect font size: {font.size}"
        assert font.bold, "Header must be bold!"
        assert fill.start_color.rgb == "00EDF2F7", f"Incorrect background: {fill.start_color.rgb}"
        
        # 6. Styling Check (Data row)
        data_cell = ws.cell(row=2, column=3) # '시가' cell
        d_font = data_cell.font
        d_align = data_cell.alignment
        d_fmt = data_cell.number_format
        
        print("Checking data cell (C2) style:")
        print(f" - Font name: {d_font.name}")
        print(f" - Font size: {d_font.size}")
        print(f" - Alignment: {d_align.horizontal}")
        print(f" - Number format: {d_fmt}")
        
        assert d_font.name == "Malgun Gothic", f"Incorrect data font: {d_font.name}"
        assert d_font.size == 10, f"Incorrect data font size: {d_font.size}"
        assert d_align.horizontal == "right", f"Price column alignment must be right! Found: {d_align.horizontal}"
        assert d_fmt == "#,##0", f"Price column formatting mismatch! Found: {d_fmt}"
        
        # Check Ticker code format (B2)
        ticker_cell = ws.cell(row=2, column=2)
        print("Checking ticker cell (B2) style:")
        print(f" - Number format: {ticker_cell.number_format}")
        print(f" - Alignment: {ticker_cell.alignment.horizontal}")
        assert ticker_cell.number_format == "@", f"Ticker code format must be text! Found: {ticker_cell.number_format}"
        assert ticker_cell.alignment.horizontal == "center", "Ticker code must be centered"
        
        # Check source column (G2)
        source_cell = ws.cell(row=2, column=7)
        print("Checking source cell (G2) style:")
        print(f" - Value: {source_cell.value}")
        print(f" - Alignment: {source_cell.alignment.horizontal}")
        assert source_cell.value in ["Naver", "KRX", "Yahoo"], f"Invalid source value: {source_cell.value}"
        assert source_cell.alignment.horizontal == "center", "Source column must be centered"
        
    print("\nALL VERIFICATIONS PASSED SUCCESSFULLY!")

if __name__ == "__main__":
    run_test_collection()
    verify_excel_file()
