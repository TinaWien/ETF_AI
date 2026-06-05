import os
import datetime
import pandas as pd
import openpyxl

def verify_report():
    print("=== Starting Detailed Excel Verification ===")
    today_str = datetime.date.today().strftime("%Y%m%d")
    filepath = f"/Users/tina/Documents/ETF_AI/output/Stock_{today_str}.xlsx"
    
    # 1. Existence check
    if not os.path.exists(filepath):
        print(f"Error: File does not exist: {filepath}")
        return False
    print(f"File found: {filepath}")
    
    # 2. Structure check using pandas
    try:
        df = pd.read_excel(filepath, dtype={"종목코드": str})
        print(f"Total rows in dataframe: {len(df)}")
    except Exception as e:
        print(f"Error reading excel with pandas: {e}")
        return False
        
    expected_cols = [
        "종목코드", "종목명", "유형", "KRX업종", "WICS업종", "현재가", 
        "시가총액(억)", "외국인비율", "투자의견", "목표주가", 
        "52주최고", "52주최저", "PER", "PBR", "배당수익률", "배당금", "관리종목"
    ]
    if list(df.columns) != expected_cols:
        print(f"Error: Column mismatch. Found {list(df.columns)}, expected {expected_cols}")
        return False
    print("Columns structure: OK")
    
    # 3. Row count sanity check
    if len(df) < 2000 or len(df) > 3500:
        print(f"Error: Row count {len(df)} is outside the normal range (2000 ~ 3500).")
        return False
    print(f"Row count sanity check: OK ({len(df)} rows)")
    
    # 4. Type count validation (Pre-existing)
    type_counts = df["유형"].value_counts()
    print("Constituents count by Type:")
    for typ, count in type_counts.items():
        print(f" - {typ}: {count}")
        
    k200_count = type_counts.get("코스피200", 0)
    if k200_count != 199 and k200_count != 200:
        print(f"Warning/Error: KOSPI 200 count is {k200_count}, expected 199 or 200.")
    else:
        print(f"KOSPI 200 count: OK ({k200_count})")
        
    k150_count = type_counts.get("코스닥150", 0)
    if k150_count != 149 and k150_count != 150:
        print(f"Warning/Error: KOSDAQ 150 count is {k150_count}, expected 149 or 150.")
    else:
        print(f"KOSDAQ 150 count: OK ({k150_count})")
        
    # 5. Null checks (only check critical columns, as PER, PBR, consensus, WICS can be empty/Null for minor stocks)
    critical_cols = ["종목코드", "종목명", "유형", "현재가", "시가총액(억)"]
    null_counts = df[critical_cols].isnull().sum()
    if null_counts.any():
        print("Error: Found null values in critical columns:")
        print(null_counts)
        return False
    print("Critical columns null check: OK")
    
    # 6. Specific stocks check (Samsung Electronics 005930)
    samsung = df[df["종목코드"] == "005930"]
    if samsung.empty:
        print("Error: Samsung Electronics (005930) is missing from the list.")
        return False
    else:
        sam_row = samsung.iloc[0]
        if sam_row["유형"] != "코스피200":
            print(f"Error: Samsung Electronics type is {sam_row['유형']}, expected 코스피200")
            return False
            
        # Sector validations
        if sam_row["KRX업종"] != "반도체와반도체장비":
            print(f"Error: Samsung Electronics KRX Sector is {sam_row['KRX업종']}, expected '반도체와반도체장비'")
            return False
        if sam_row["WICS업종"] != "반도체와반도체장비":
            print(f"Error: Samsung Electronics WICS Sector is {sam_row['WICS업종']}, expected '반도체와반도체장비'")
            return False
            
        # Detailed metrics validations
        if sam_row["목표주가"] <= 0 or sam_row["52주최고"] <= 0 or sam_row["52주최저"] <= 0:
            print(f"Error: Invalid 52-week or target price values for Samsung. Row: {sam_row.to_dict()}")
            return False
            
        # Ratios validations
        if pd.isnull(sam_row["PER"]) or pd.isnull(sam_row["PBR"]) or pd.isnull(sam_row["배당수익률"]):
            print(f"Error: Missing ratios (PER/PBR/Yield) for Samsung. Row: {sam_row.to_dict()}")
            return False
            
        print("Samsung Electronics (005930) sector, mapping, metrics and values: OK")
        
    # 7. Excel cell formatting and styling checks
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
        # Grid lines check
        if not ws.views.sheetView[0].showGridLines:
            print("Error: Sheet grid lines are disabled.")
            return False
        print("Grid lines: OK (Enabled)")
        
        # Header check (Row 1)
        for col_idx in range(1, len(expected_cols) + 1):
            cell = ws.cell(row=1, column=col_idx)
            if cell.font.name != "Malgun Gothic":
                print(f"Error: Header font name is {cell.font.name}, expected 'Malgun Gothic'")
                return False
            if not cell.font.bold:
                print("Error: Header font is not Bold.")
                return False
            fill_color = cell.fill.start_color.rgb
            if fill_color not in ["00D9E1F2", "FFD9E1F2"]:
                print(f"Error: Header fill color is {fill_color}, expected D9E1F2.")
                return False
        print("Header fonts & colors: OK")
        
        # Data format check (Row 2, Samsung should be first or second row as it has highest market cap)
        # We will find the Samsung row in excel
        sam_row_idx = None
        for r in range(2, 10):
            if ws.cell(row=r, column=1).value == "005930":
                sam_row_idx = r
                break
                
        if not sam_row_idx:
            print("Error: Could not find Samsung row in top market cap items in Excel.")
            return False
            
        # Code format
        c_code = ws.cell(row=sam_row_idx, column=1)
        if c_code.number_format != "@":
            print(f"Error: Stock Code format is {c_code.number_format}, expected '@'")
            return False
            
        # WICS Sector Alignment
        c_wics = ws.cell(row=sam_row_idx, column=5)
        if c_wics.alignment.horizontal != "center":
            print(f"Error: WICS Sector alignment is {c_wics.alignment.horizontal}, expected 'center'")
            return False
            
        # Target Price format
        c_target = ws.cell(row=sam_row_idx, column=10)
        if c_target.number_format != "#,##0":
            print(f"Error: Target price format is {c_target.number_format}, expected '#,##0'")
            return False
            
        # PER format
        c_per = ws.cell(row=sam_row_idx, column=13)
        if c_per.number_format != "0.00":
            print(f"Error: PER format is {c_per.number_format}, expected '0.00'")
            return False
            
        # Dividend Yield format
        c_yield = ws.cell(row=sam_row_idx, column=15)
        if c_yield.number_format != "0.00%":
            print(f"Error: Yield format is {c_yield.number_format}, expected '0.00%'")
            return False
            
        print("Excel cell alignments & number formatting: OK")
        
    except Exception as e:
        print(f"Error during openpyxl styling check: {e}")
        return False
        
    print("=== All Verifications Passed! ===")
    return True

if __name__ == "__main__":
    import sys
    success = verify_report()
    if not success:
        sys.exit(1)
